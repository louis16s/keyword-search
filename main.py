import os
import time
import configparser
import requests
import warnings
from multiprocessing import Pool, cpu_count
from openpyxl import load_workbook, Workbook
import xlrd
from tqdm import tqdm
from rich.console import Console
from rich.panel import Panel
from rich.prompt import Prompt
from rich.table import Table
import shutil
from packaging import version

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
console = Console()
VERSION = "2.2"
REPO_API = "https://api.github.com/repos/louis16s/keyword-search/releases/latest"


def read_keywords_from_excel(excel_path):
    keywords = []
    try:
        if excel_path.endswith('.xlsx'):
            workbook = load_workbook(filename=excel_path, read_only=True, data_only=True)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
                keyword = row[0]
                if keyword:
                    keywords.append(str(keyword).strip())
        elif excel_path.endswith('.xls'):
            workbook = xlrd.open_workbook(excel_path)
            sheet = workbook.sheet_by_index(0)
            for row_idx in range(sheet.nrows):
                keyword = sheet.cell_value(row_idx, 0)
                if keyword:
                    keywords.append(str(keyword).strip())
        else:
            console.print(f"[red]不支持的关键词文件类型：{excel_path}[/red]")
    except Exception as e:
        console.print(f"[red]读取关键词文件失败：{e}[/red]")
    return keywords


def process_single_file(args):
    file_path, keywords = args
    matched_rows = []
    try:
        filename = os.path.basename(file_path)
        if file_path.endswith('.xlsx'):
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    if row and any(keyword in str(cell) for cell in row if cell for keyword in keywords):
                        matched_rows.append([filename] + [str(cell) if cell is not None else '' for cell in row])
        elif file_path.endswith('.xls'):
            wb = xlrd.open_workbook(file_path)
            for sheet in wb.sheets():
                for row_idx in range(sheet.nrows):
                    row = sheet.row(row_idx)
                    if any(keyword in str(cell.value) for cell in row for keyword in keywords):
                        matched_rows.append([filename] + [str(cell.value) for cell in row])
    except Exception as e:
        console.print(f"[red]错误读取文件 {file_path}：{e}[/red]")
    return matched_rows


def search_keywords_parallel(keywords, folder_path, output_excel_path):
    all_excel_files = []
    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.endswith(('.xlsx', '.xls')):
                all_excel_files.append(os.path.join(root, file))

    console.print(Panel(f"共找到 {len(all_excel_files)} 个 Excel 文件，开始搜索...", title="文件加载完毕"), justify="center")

    results = []
    with Pool(processes=cpu_count()) as pool:
        for file_result in tqdm(pool.imap_unordered(process_single_file, [(file, keywords) for file in all_excel_files]),
                                total=len(all_excel_files), desc="搜索中"):
            if file_result:
                results.extend(file_result)

    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "搜索结果"
    output_ws.append(["文件名", "匹配行内容"])

    for row in results:
        output_ws.append(row)

    output_wb.save(output_excel_path)
    console.print(Panel(f"搜索完成！结果已保存至 [green]{output_excel_path}[/green]", title="完成"), justify="center")
    input("\n按回车返回菜单...")


def file_path_read():
    config = configparser.ConfigParser()
    if not os.path.exists('config.ini'):
        with open('config.ini', 'w', encoding='utf-8') as file:
            file.write(
                '[Settings]\n'
                'search_directory = 输入你要搜索的文件夹路径\n'
                'excel_file_path = 输入包含关键词的Excel路径\n'
            )
        console.print(Panel("[bold yellow]配置文件已创建：config.ini\n请编辑后重新运行本程序。[/bold yellow]", title="初始化完成"), justify="center")
        time.sleep(5)
        exit(0)

    config.read('config.ini', encoding='utf-8')
    folder_to_search = config['Settings']['search_directory'].strip()
    excel_file_path = config['Settings']['excel_file_path'].strip()
    output_excel_path = "搜索结果.xlsx"
    return folder_to_search, excel_file_path, output_excel_path


def modify_config():
    config = configparser.ConfigParser()
    config.read('config.ini', encoding='utf-8')

    if 'Settings' not in config:
        config['Settings'] = {'search_directory': '', 'excel_file_path': ''}

    while True:
        settings = list(config['Settings'].items())
        console.print("\n[bold green]当前配置：[/bold green]")
        table = Table(show_header=True, header_style="bold cyan")
        table.add_column("序号", style="dim", width=6)
        table.add_column("配置项", width=20)
        table.add_column("当前值")

        table.add_row("0", "[dim]返回上一页[/dim]", "")
        for idx, (key, val) in enumerate(settings, 1):
            table.add_row(str(idx), key, val)
        console.print(table)

        choice = Prompt.ask("\n请输入要修改的配置项序号（输入 0 返回）", choices=[str(i) for i in range(0, len(settings)+1)])
        if choice == "0":
            break

        key_to_modify = settings[int(choice)-1][0]
        new_val = Prompt.ask(f"请输入新的值 (原值: {config['Settings'][key_to_modify]})")
        config['Settings'][key_to_modify] = new_val
        with open('config.ini', 'w', encoding='utf-8') as f:
            config.write(f)
        console.print("[green]配置修改完成！[/green]\n")


def check_for_updates():
    console.print("\n[bold cyan]正在检查更新...[/bold cyan]", justify="center")
    try:
        response = requests.get(REPO_API, timeout=5)
        if response.status_code == 200:
            latest_release = response.json()
            latest_version = latest_release["tag_name"].lstrip("v")
            download_url = None
            asset_name = None

            for asset in latest_release.get("assets", []):
                if asset["name"].endswith(".zip") or asset["name"].endswith(".exe"):
                    download_url = asset["browser_download_url"]
                    asset_name = asset["name"]
                    break

            if version.parse(latest_version) > version.parse(VERSION):
                console.print(Panel(
                    f"发现新版本：v{latest_version}（当前版本：v{VERSION}）\n"
                    f"正在下载更新文件：{asset_name or '未命名'}",
                    title="更新可用"), justify="center")

                if download_url:
                    with requests.get(download_url, stream=True) as r:
                        r.raise_for_status()
                        with open(asset_name, 'wb') as f:
                            shutil.copyfileobj(r.raw, f)
                    console.print(Panel(
                        f"[green]下载完成！文件已保存为：{asset_name}[/green]\n"
                        "请手动关闭程序并运行新版本。",
                        title="更新完成"), justify="center")
                else:
                    console.print("[red]未找到有效的下载链接。请手动前往页面下载。[/red]", justify="center")
            else:
                console.print(Panel("你已经是最新版本啦~", title="暂无更新"), justify="center")
        else:
            console.print("[red]无法检查更新，请稍后再试。[/red]", justify="center")
    except Exception as e:
        console.print(f"[red]更新检查失败：{e}[/red]", justify="center")
    input("\n按回车返回菜单...")


def show_menu():
    while True:
        os.system('cls' if os.name == 'nt' else 'clear')
        console.print(Panel.fit("[bold cyan]Excel 关键词搜索工具[/bold cyan] " +VERSION+ " \n power by louis16s", title="主菜单", subtitle="请选择操作"), justify="center")

        menu_table = Table(show_header=False)
        menu_table.add_row("[bold] 1. [/bold]", "   运行关键词搜索   ")
        menu_table.add_row("[bold] 2. [/bold]", "    修改配置文件    ")
        menu_table.add_row("[bold] 3. [/bold]", "      检查更新     ")
        menu_table.add_row("[bold] 4. [/bold]", "      退出程序     ")
        console.print(menu_table, justify="center")

        choice = Prompt.ask("\n请输入你的选择", choices=["1", "2", "3", "4"])
        if choice == "1":
            folder_to_search, excel_file_path, output_excel_path = file_path_read()
            console.print(f"[cyan]正在读取关键词文件：{excel_file_path}[/cyan]", justify="center")
            keywords = read_keywords_from_excel(excel_file_path)
            search_keywords_parallel(keywords, folder_to_search, output_excel_path)
        elif choice == "2":
            modify_config()
        elif choice == "3":
            check_for_updates()
        elif choice == "4":
            console.print("\n[bold green]感谢使用，再见！[/bold green]", justify="center")
            break


if __name__ == "__main__":
    show_menu()
